"""セミナー申込フォーム 保存用ダミーデータ生成スクリプト。

`walkthrough_seminar_demo.md` のデモフォーム（セミナー申込フォーム）の保存先
スプレッドシートと**同じレイアウト**で大量のダミー回答を生成する。

GAS 側 (Sheets_buildOrderFromSchema_ / Sheets_initializeHeaders_ /
Sheets_upsertRecordById_) と同じく、11 行ヘッダー (NFB_HEADER_DEPTH=11) +
12 行目以降データ (NFB_DATA_START_ROW=12)、固定メタ列 + 動的列のレイアウトを
再現する。選択肢系（ドロップダウン/ラジオ/チェックボックス）は選択肢ごとに 1 列で、
選んだ列にマーカー "●" を立てる（gas/sheetsHeaders.gs:187 と同じ仕様）。

実際の保存先シート（読み取りで確認した 23 列）の列順:
  id, No., createdAt, modifiedAt, deletedAt, createdBy, modifiedBy, deletedBy, pid,
  申込者氏名,
  所属区分/行政, 所属区分/企業, 所属区分/学生, 所属区分/その他,
  参加形態/会場参加, 参加形態/オンライン参加,
  参加形態/会場参加/希望座席/前方, …/中央, …/後方,
  参加形態/オンライン参加/接続方法/PC, …/スマートフォン, …/タブレット,
  参加希望日

使い方:
  python scripts/generate_seminar_records.py --count 120 --out form_data/seminar_dummy.xlsx
  # 生成された xlsx の 12 行目以降（または同時出力の .tsv）を、保存先スプレッドシートの
  # 12 行目以降に貼り付けると、次回同期時にアプリへ取り込まれる。

注意:
  - createdAt / modifiedAt は unix ミリ秒（数値）。GAS 側 sheetsDatetime.gs が
    UNIX_MS_THRESHOLD(=1e11) 以上の数値を日時として解釈する。
  - 参加希望日（日付フィールド）は "YYYY/MM/DD" 文字列で書く
    （generate_camera_records.py の date 列と同じ方針）。挙動が合わない場合は
    DATE_AS_UNIX_MS=True にして unix ミリ秒で書く。
"""
from __future__ import annotations

import argparse
import random
import secrets
import time
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl が必要です: pip install openpyxl") from exc

# --- 定数（GAS constants.gs と一致させる） --------------------------------
NFB_HEADER_DEPTH = 11
NFB_DATA_START_ROW = NFB_HEADER_DEPTH + 1  # 12
CHOICE_MARK = "●"
DATE_AS_UNIX_MS = False  # True にすると参加希望日も unix ミリ秒で書く

# 固定メタ列（現行は pid を含む 9 列）
FIXED_META = ["id", "No.", "createdAt", "modifiedAt", "deletedAt",
              "createdBy", "modifiedBy", "deletedBy", "pid"]

# 動的列（スラッシュ区切りのパス。実シートの列順そのまま）
DYNAMIC_COLUMNS = [
    "申込者氏名",
    "所属区分/行政", "所属区分/企業", "所属区分/学生", "所属区分/その他",
    "参加形態/会場参加", "参加形態/オンライン参加",
    "参加形態/会場参加/希望座席/前方",
    "参加形態/会場参加/希望座席/中央",
    "参加形態/会場参加/希望座席/後方",
    "参加形態/オンライン参加/接続方法/PC",
    "参加形態/オンライン参加/接続方法/スマートフォン",
    "参加形態/オンライン参加/接続方法/タブレット",
    "参加希望日",
]

ALL_COLUMNS = FIXED_META + DYNAMIC_COLUMNS

# --- 選択肢マスタ ---------------------------------------------------------
AFFIL = ["行政", "企業", "学生", "その他"]
AFFIL_WEIGHTS = [4, 3, 2, 1]
MODE = ["会場参加", "オンライン参加"]
MODE_WEIGHTS = [3, 2]
SEAT = ["前方", "中央", "後方"]
CONN = ["PC", "スマートフォン", "タブレット"]

SEI = ["山田", "佐藤", "鈴木", "田中", "高橋", "伊藤", "中村", "渡辺", "小林", "加藤",
       "吉田", "山本", "斎藤", "松本", "井上", "木村", "林", "清水", "森", "池田"]
MEI = ["太郎", "花子", "一郎", "美咲", "健", "葵", "大輔", "七海", "翔", "結衣",
       "陽菜", "拓也", "彩", "直樹", "さくら", "悠斗", "茜", "亮", "莉子", "蓮"]

ULID_ALPHABET = "0123456789ABCDEFGHJKMNPQRSTVWXYZ"
URLSAFE = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"


# --- レコード ID 生成（r_<ULID26>_<base64url8>） --------------------------
def encode_ulid(ts_ms: int) -> str:
    time_chars = []
    t = ts_ms
    for _ in range(10):
        time_chars.append(ULID_ALPHABET[t % 32])
        t //= 32
    rand_chars = [secrets.choice(ULID_ALPHABET) for _ in range(16)]
    return "".join(reversed(time_chars)) + "".join(rand_chars)


def generate_record_id(ts_ms: int, used: set[str]) -> str:
    while True:
        rid = "r_" + encode_ulid(ts_ms) + "_" + "".join(secrets.choice(URLSAFE) for _ in range(8))
        if rid not in used:
            used.add(rid)
            return rid


# --- ヘッダー行列（11 行 × N 列） -----------------------------------------
def build_header_matrix() -> list[list[str]]:
    matrix = [["" for _ in ALL_COLUMNS] for _ in range(NFB_HEADER_DEPTH)]
    for col_idx, key in enumerate(ALL_COLUMNS):
        segments = key.split("/")
        for row_idx, seg in enumerate(segments):
            if row_idx < NFB_HEADER_DEPTH:
                matrix[row_idx][col_idx] = seg
    return matrix


# --- 1 レコード分の値（列キー → 値） --------------------------------------
def build_record_values(i: int, now_ms: int, email: str) -> dict[str, object]:
    affil = random.choices(AFFIL, weights=AFFIL_WEIGHTS, k=1)[0]
    mode = random.choices(MODE, weights=MODE_WEIGHTS, k=1)[0]
    seat = random.choice(SEAT) if mode == "会場参加" else None
    conn = random.choice(CONN) if mode == "オンライン参加" else None
    name = random.choice(SEI) + random.choice(MEI)
    want_day = date(2026, 7, 1) + timedelta(days=random.randint(0, 45))

    if DATE_AS_UNIX_MS:
        want_val: object = int(datetime(want_day.year, want_day.month, want_day.day).timestamp() * 1000)
    else:
        want_val = want_day.strftime("%Y/%m/%d")

    values: dict[str, object] = {k: "" for k in ALL_COLUMNS}
    values["申込者氏名"] = name
    for opt in AFFIL:
        values[f"所属区分/{opt}"] = CHOICE_MARK if opt == affil else ""
    for opt in MODE:
        values[f"参加形態/{opt}"] = CHOICE_MARK if opt == mode else ""
    for opt in SEAT:
        values[f"参加形態/会場参加/希望座席/{opt}"] = CHOICE_MARK if seat == opt else ""
    for opt in CONN:
        values[f"参加形態/オンライン参加/接続方法/{opt}"] = CHOICE_MARK if conn == opt else ""
    values["参加希望日"] = want_val
    return values


def build_meta(i: int, now_ms: int, email: str, used_ids: set[str]) -> dict[str, object]:
    # createdAt を少しずつずらして時系列っぽくする
    created = now_ms - (count_offset := (random.randint(0, 60 * 24 * 60 * 60 * 1000)))
    return {
        "id": generate_record_id(created, used_ids),
        "No.": i + 1,
        "createdAt": created,
        "modifiedAt": created,
        "deletedAt": "",
        "createdBy": email,
        "modifiedBy": email,
        "deletedBy": "",
        "pid": "",
    }


# --- 出力 -----------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description="セミナー申込フォーム 保存用ダミーデータ生成")
    ap.add_argument("--count", type=int, default=120, help="生成件数（既定 120）")
    ap.add_argument("--out", type=Path, default=Path("seminar_dummy.xlsx"), help="出力 xlsx パス")
    ap.add_argument("--tsv", type=Path, default=None, help="貼り付け用 TSV（データ行のみ）の出力パス。未指定なら out と同名 .tsv")
    ap.add_argument("--email", type=str, default="gk3t.mtmr@gmail.com", help="createdBy/modifiedBy に入れるメール")
    ap.add_argument("--sheet-name", type=str, default="Data", help="xlsx のシート名")
    ap.add_argument("--seed", type=int, default=None, help="乱数シード（再現用）")
    args = ap.parse_args()

    if args.seed is not None:
        random.seed(args.seed)

    now_ms = int(time.time() * 1000)
    used_ids: set[str] = set()
    header = build_header_matrix()

    # データ行（メタ + 動的）
    rows: list[list[object]] = []
    for i in range(args.count):
        meta = build_meta(i, now_ms, args.email, used_ids)
        vals = build_record_values(i, now_ms, args.email)
        merged = {**vals, **meta}  # メタ列を優先（vals は ALL_COLUMNS を空初期化しているため）
        rows.append([merged[k] for k in ALL_COLUMNS])

    # createdAt 昇順 + No. 振り直し
    rows.sort(key=lambda r: r[ALL_COLUMNS.index("createdAt")])
    for n, r in enumerate(rows):
        r[ALL_COLUMNS.index("No.")] = n + 1

    # xlsx 出力（11 行ヘッダー + 12 行目以降データ）
    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet_name
    for r in range(NFB_HEADER_DEPTH):
        for c in range(len(ALL_COLUMNS)):
            ws.cell(row=r + 1, column=c + 1, value=header[r][c])
    for ri, row in enumerate(rows):
        for c, v in enumerate(row):
            ws.cell(row=NFB_DATA_START_ROW + ri, column=c + 1, value=v)
    for c in range(len(ALL_COLUMNS)):
        ws.column_dimensions[get_column_letter(c + 1)].width = 16
    ws.freeze_panes = ws.cell(row=NFB_DATA_START_ROW, column=1)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    # 貼り付け用 TSV（データ行のみ。スプレッドシート 12 行目以降に貼る）
    tsv_path = args.tsv or args.out.with_suffix(".tsv")
    with open(tsv_path, "w", encoding="utf-8", newline="") as f:
        for row in rows:
            f.write("\t".join("" if v == "" else str(v) for v in row) + "\n")

    print(f"生成: {len(rows)} 件")
    print(f"  xlsx: {args.out}")
    print(f"  tsv : {tsv_path}（保存先スプレッドシートの 12 行目以降に貼り付け）")


if __name__ == "__main__":
    main()
