"""出力 xlsx を読み返してサンプルレコードを検証する確認スクリプト。"""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

DST = Path(r"C:\Users\sa11882\nested_form_builder\form_data\自動カメラ設置状況_保存用.xlsx")

wb = load_workbook(DST, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# ヘッダー (11 行) を読み取り
n_cols = ws.max_column
headers = []
for c in range(1, n_cols + 1):
    path = []
    for r in range(1, 12):
        v = ws.cell(r, c).value
        if v is None or str(v).strip() == "":
            break
        path.append(str(v).strip())
    headers.append(path)

print("\n=== Headers (path per column) ===")
for i, p in enumerate(headers):
    print(f"  col {i + 1:2d}: {' | '.join(p)}")

# データ行
data_start = 12
data_rows = ws.max_row - data_start + 1
print(f"\n=== Data row count: {data_rows} ===")

key_to_col = {"|".join(p): i + 1 for i, p in enumerate(headers)}


def get(row_num: int, key: str):
    c = key_to_col.get(key)
    if c is None:
        return None
    return ws.cell(row_num, c).value


# 全レコードの番号を表示
print("\n=== All records (No. / 番号 / 鍵 / シリアル番号) ===")
for r in range(data_start, ws.max_row + 1):
    no = get(r, "No.")
    ban = get(r, "番号")
    kagi = get(r, "鍵")
    ser = get(r, "シリアル番号")
    print(f"  row {r}: No.={no}, 番号={ban}, 鍵={kagi}, シリアル={ser}")

# サンプルレコード詳細
print("\n=== Sample records ===")
for target in ["H01", "H02", "H03", "H07", "H31", "H35", "H36"]:
    for r in range(data_start, ws.max_row + 1):
        if str(get(r, "番号") or "").strip() == target:
            print(f"\n--- {target} (row {r}) ---")
            for key in [
                "id", "No.", "createdAt", "modifiedAt", "deletedAt",
                "createdBy", "modifiedBy", "deletedBy",
                "番号", "鍵", "IMEI", "SIMNo.", "シリアル番号",
                "設置場所①", "設置場所①|設置開始日", "設置場所①|設置終了日",
                "設置場所①|座標", "設置場所①|備考",
                "設置場所②", "設置場所②|設置開始日", "設置場所②|設置終了日",
                "設置場所②|座標", "設置場所②|備考",
                "設置場所③", "設置場所③|設置開始日", "設置場所③|設置終了日",
                "設置場所③|座標", "設置場所③|備考",
                "全体備考",
            ]:
                v = get(r, key)
                if v not in (None, ""):
                    print(f"    {key}: {v!r}")
            break

# id がユニークか
ids = [get(r, "id") for r in range(data_start, ws.max_row + 1)]
print(f"\n=== ID uniqueness ===")
print(f"  unique ids: {len(set(ids))} / {len(ids)}")
print(f"  sample id : {ids[0]}")
