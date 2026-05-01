"""R8●ひぐまとめ.xlsx カメラシート → 自動カメラ設置状況フォーム保存用 xlsx 変換。

GAS 側 (Sheets_buildOrderFromSchema_ / Sheets_initializeHeaders_ /
Sheets_upsertRecordById_) と同じレイアウト (NFB_HEADER_DEPTH=11 行ヘッダー +
NFB_DATA_START_ROW=12 行目以降データ、固定メタ列 8 + 動的列) を再現する。
"""
from __future__ import annotations

import base64
import json
import os
import random
import re
import secrets
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(r"C:\Users\sa11882\nested_form_builder")
SRC_XLSX = REPO_ROOT / "form_data" / "R8●ひぐまとめ.xlsx"
SRC_JSON = REPO_ROOT / "form_data" / "自動カメラ設置状況.json"
DST_XLSX = REPO_ROOT / "form_data" / "自動カメラ設置状況_保存用.xlsx"

NFB_HEADER_DEPTH = 11
NFB_DATA_START_ROW = NFB_HEADER_DEPTH + 1  # 12
NFB_FIXED_HEADER_PATHS = [
    ["id"], ["No."], ["createdAt"], ["modifiedAt"], ["deletedAt"],
    ["createdBy"], ["modifiedBy"], ["deletedBy"],
]
USER_EMAIL = "gk3t.mtmr@gmail.com"
ULID_ALPHABET = "0123456789ABCDEFGHJKMNPQRSTVWXYZ"

SINGLE_VALUE_TYPES = {
    "text", "textarea", "number", "regex", "date", "time",
    "url", "userName", "email", "phone", "fileUpload", "substitution",
}


# --- ヘッダー / 正規化 (GAS Sheets_normalize* と等価) ----------------------

def normalize_header_segment(segment) -> str:
    if segment is None:
        return ""
    s = str(segment)
    s = re.sub(r"\r\n?", "\n", s)
    return s.strip()


def normalize_header_path(path) -> list[str]:
    out: list[str] = []
    if not isinstance(path, (list, tuple)):
        return out
    for i, seg in enumerate(path):
        if i >= NFB_HEADER_DEPTH:
            break
        n = normalize_header_segment(seg)
        if not n:
            break
        out.append(n)
    return out


def path_key(path) -> str:
    return "|".join(normalize_header_path(path))


def normalize_header_key(key) -> str:
    if key is None:
        return ""
    return path_key(str(key).split("|"))


# --- スキーマ走査 (GAS nfbTraverseSchema_ / Sheets_buildOrderFromSchema_) ---

def field_segment_with_fallback(field, index_trail) -> str:
    label = normalize_header_segment(field.get("label") if isinstance(field, dict) else None)
    if label:
        return label
    t = ""
    if isinstance(field, dict) and field.get("type") is not None:
        t = str(field.get("type")).strip()
    return f"質問 {'.'.join(str(x) for x in index_trail)} ({t or 'unknown'})"


def build_order_from_schema(schema) -> list[str]:
    order: list[str] = []
    seen: set[str] = set()

    def append_key(key: str):
        n = normalize_header_key(key)
        if not n or n in seen:
            return
        seen.add(n)
        order.append(n)

    def walk(nodes, path_segments, depth, index_trail):
        if not isinstance(nodes, list):
            return
        for i, field in enumerate(nodes):
            if field is None:
                continue
            current_index_trail = list(index_trail) + [i + 1]
            seg = field_segment_with_fallback(field, current_index_trail)
            if seg is None:
                continue
            current_path = list(path_segments) + [seg]
            t = str(field.get("type", "")).strip() if isinstance(field, dict) else ""
            base_key = "|".join(current_path)

            if t in ("checkboxes", "radio", "select"):
                opts = field.get("options") if isinstance(field, dict) else None
                if isinstance(opts, list):
                    for opt in opts:
                        opt_label = normalize_header_segment(opt.get("label") if isinstance(opt, dict) else None)
                        append_key(base_key + "|" + opt_label if opt_label else base_key + "|")
            elif t != "message" and t in SINGLE_VALUE_TYPES:
                append_key(base_key)

            children = field.get("children") if isinstance(field, dict) else None
            if isinstance(children, list) and children:
                walk(children, current_path, depth + 1, current_index_trail)

    walk(schema if isinstance(schema, list) else [], [], 1, [])
    return order


# --- ヘッダー matrix (NFB_HEADER_DEPTH 行 × N 列) ---------------------------

def build_header_matrix(order_keys: list[str]) -> list[list[str]]:
    """[NFB_HEADER_DEPTH 行][N 列] の 2 次元配列を返す。"""
    paths: list[list[str]] = []
    seen: set[str] = set()

    # 固定列を先頭に
    for p in NFB_FIXED_HEADER_PATHS:
        np_ = normalize_header_path(p)
        if not np_:
            continue
        k = path_key(np_)
        if k not in seen:
            paths.append(np_)
            seen.add(k)

    # 動的列
    for k_raw in order_keys:
        parts = normalize_header_path(str(k_raw or "").split("|"))
        if not parts:
            continue
        k = path_key(parts)
        if k not in seen:
            paths.append(parts)
            seen.add(k)

    matrix = [["" for _ in paths] for _ in range(NFB_HEADER_DEPTH)]
    for col, p in enumerate(paths):
        for row in range(NFB_HEADER_DEPTH):
            matrix[row][col] = p[row] if row < len(p) else ""
    return matrix, paths


# --- ID 生成 (GAS Nfb_generateRecordId_ と互換: r_<ULID26>_<base64url8>) ----

def encode_ulid_time(unix_ms: int) -> str:
    if unix_ms < 0:
        unix_ms = 0
    chars = []
    v = unix_ms
    for _ in range(10):
        chars.append(ULID_ALPHABET[v % 32])
        v //= 32
    return "".join(reversed(chars))


def encode_ulid_random_part() -> str:
    raw = secrets.token_bytes(10)
    buf = 0
    bits = 0
    out = []
    for b in raw:
        buf = (buf << 8) | b
        bits += 8
        while bits >= 5:
            out.append(ULID_ALPHABET[(buf >> (bits - 5)) & 31])
            bits -= 5
            buf &= (1 << bits) - 1 if bits > 0 else 0
    if bits > 0:
        out.append(ULID_ALPHABET[(buf << (5 - bits)) & 31])
    return "".join(out)[:16]


def generate_record_id() -> str:
    ulid = encode_ulid_time(int(time.time() * 1000)) + encode_ulid_random_part()
    rand_part = base64.urlsafe_b64encode(secrets.token_bytes(6)).decode("ascii").rstrip("=")[:8]
    return f"r_{ulid}_{rand_part}"


# --- Excel 読み取り + マージ伝播 ------------------------------------------

def read_camera_sheet_with_merges(xlsx_path: Path) -> tuple[list[list], int, int]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["カメラ"]
    max_row = ws.max_row
    max_col = ws.max_column
    data = [[ws.cell(r, c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
    for m in ws.merged_cells.ranges:
        v = ws.cell(m.min_row, m.min_col).value
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                data[r - 1][c - 1] = v
    wb.close()
    return data, max_row, max_col


# --- 値の正規化 -----------------------------------------------------------

DATE_DOT_RE = re.compile(r"^\s*(\d{4})\.(\d{1,2})\.(\d{1,2})\s*$")
COORD_RE = re.compile(r"(\d+\.\d+)\s*[,、]\s*(\d+\.\d+)")


def to_date_string(v) -> str:
    """Excel シリアル / datetime / 'yyyy.M.D' → 'yyyy/MM/dd'。解釈不能なら空。"""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=int(v))
            return d.strftime("%Y/%m/%d")
        except Exception:
            return ""
    s = str(v).strip()
    m = DATE_DOT_RE.match(s)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), m.group(3).zfill(2)
        return f"{y}/{mo}/{d}"
    return ""


def imei_sim_to_str(v, expected_digits: int) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, bool):
        return ""
    if isinstance(v, (int, float)):
        try:
            n = int(round(float(v)))
        except Exception:
            return ""
        s = str(n)
        return s if len(s) == expected_digits else ""
    s = str(v).strip()
    s_digits = re.sub(r"\D", "", s)
    return s_digits if len(s_digits) == expected_digits else ""


def normalize_kagi(s) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    # [なし] / 【なし】 / [な し] / なし のみ → 空
    if re.fullmatch(r"[\[【［\(]?\s*な\s*し\s*[\]】］\)]?", t):
        return ""
    return t


def parse_number_kagi(a_cell) -> tuple[str, str]:
    """A 列パース。Excel 上では番号と鍵が改行で区切られていることが多い:
       'H01\\n[札02]' → ('H01', '札02')
       'H22\\n[★22]' → ('H22', '★22')
       'H35\\n【なし】' → ('H35', '') - normalize_kagi で「なし」を空に
       'H31' → ('H31', '') - 鍵記述なし
       'H25\\n鍵なし' → ('H25', '') - 括弧なしで '鍵なし' のみ
    """
    if a_cell is None:
        return "", ""
    s = str(a_cell).strip()
    m = re.match(r"^(.+?)\s*[\[【［\(]\s*(.*?)\s*[\]】］\)]\s*$", s, re.DOTALL)
    if m:
        return m.group(1).strip(), normalize_kagi(m.group(2))
    return s.split("\n")[0].strip(), ""


def split_coord_and_remark(remark) -> tuple[str, str]:
    """備考から座標 'lat,lng' を抽出。残りを備考として返す。"""
    if remark is None:
        return "", ""
    text = str(remark).strip()
    if not text:
        return "", ""
    m = COORD_RE.search(text)
    if not m:
        return "", text
    lat, lng = m.group(1), m.group(2)
    coord = f"{lat},{lng}"
    rest = text[: m.start()] + text[m.end():]
    rest = re.sub(r"^[\s　,、]+|[\s　,、]+$", "", rest)
    return coord, rest


def build_location_text(ku, basho, hosoku) -> str:
    ku_s = str(ku).strip() if ku not in (None, "") else ""
    f_s = str(basho).strip() if basho not in (None, "") else ""
    g_s = str(hosoku).strip() if hosoku not in (None, "") else ""
    if not (ku_s or f_s or g_s):
        return ""
    head = f"{ku_s}）" if ku_s else ""
    body = f_s
    tail = ""
    if g_s and g_s != f_s:
        if not (f_s.endswith("）") and g_s in f_s):
            tail = f"（{g_s}）"
    return f"{head}{body}{tail}"


# --- メインビルダー -------------------------------------------------------

# Excel 列インデックス (1-based)
COL_NUMBER = 1   # A: 番号[鍵]
COL_PERIOD_START = 2  # B
COL_PERIOD_TILDE = 3  # C: ～
COL_PERIOD_END = 4    # D
COL_KU = 5            # E
COL_LOCATION = 6      # F
COL_LOCATION_SUB = 7  # G
COL_REMARK = 8        # H
COL_IMEI = 11         # K
COL_SIM = 12          # L
COL_SERIAL = 13       # M


def collect_records(rows: list[list]) -> list[dict]:
    """マージ伝播済み rows (1-based 行 → 0-based index) からカメラ単位レコードを構築。"""
    HEADER_END_ROW = 4  # 4 行目までヘッダー
    groups: dict[str, list[int]] = {}
    order_serial: list[str] = []

    for ridx in range(HEADER_END_ROW, len(rows)):
        row = rows[ridx]
        serial = row[COL_SERIAL - 1] if len(row) >= COL_SERIAL else None
        if serial is None or str(serial).strip() == "":
            continue
        key = str(serial).strip()
        if key not in groups:
            groups[key] = []
            order_serial.append(key)
        groups[key].append(ridx)

    records = []
    for serial in order_serial:
        ridxs = groups[serial]
        first = rows[ridxs[0]]
        number, kagi = parse_number_kagi(first[COL_NUMBER - 1])
        imei = imei_sim_to_str(first[COL_IMEI - 1], 15) if len(first) >= COL_IMEI else ""
        sim = imei_sim_to_str(first[COL_SIM - 1], 14) if len(first) >= COL_SIM else ""
        ser_str = str(serial).strip()

        entries = []
        for ridx in ridxs:
            row = rows[ridx]
            ku = row[COL_KU - 1] if len(row) >= COL_KU else None
            basho = row[COL_LOCATION - 1] if len(row) >= COL_LOCATION else None
            hosoku = row[COL_LOCATION_SUB - 1] if len(row) >= COL_LOCATION_SUB else None
            remark_raw = row[COL_REMARK - 1] if len(row) >= COL_REMARK else None
            tilde = row[COL_PERIOD_TILDE - 1] if len(row) >= COL_PERIOD_TILDE else None
            start = row[COL_PERIOD_START - 1] if len(row) >= COL_PERIOD_START else None
            end = row[COL_PERIOD_END - 1] if len(row) >= COL_PERIOD_END else None

            location_text = build_location_text(ku, basho, hosoku)
            start_s = to_date_string(start)
            tilde_present = tilde is not None and str(tilde).strip() != ""
            end_s = to_date_string(end) if tilde_present else ""
            coord, rest_remark = split_coord_and_remark(remark_raw)

            # 完全に空のエントリーはスキップ
            if not any([location_text, start_s, end_s, coord, rest_remark]):
                continue

            entries.append({
                "location": location_text,
                "start": start_s,
                "end": end_s,
                "coord": coord,
                "remark": rest_remark,
            })

        records.append({
            "serial": ser_str,
            "number": number,
            "kagi": kagi,
            "imei": imei,
            "sim": sim,
            "entries": entries,
        })

    # 番号昇順にソート (数値混在対応: 英字+数字)
    def number_sort_key(rec):
        n = rec["number"] or ""
        m = re.match(r"^([A-Za-z]+)(\d+)$", n)
        if m:
            return (m.group(1), int(m.group(2)))
        return (n, 0)

    records.sort(key=number_sort_key)
    return records


def build_dynamic_responses(rec: dict) -> dict:
    """動的列キー → 値 のマップ。"""
    out = {}
    out["番号"] = rec["number"]
    out["鍵"] = rec["kagi"]
    out["IMEI"] = rec["imei"]
    out["SIMNo."] = rec["sim"]
    out["シリアル番号"] = rec["serial"]

    entries = rec["entries"]
    labels = ["設置場所①", "設置場所②", "設置場所③"]
    for i, label in enumerate(labels):
        if i < len(entries):
            e = entries[i]
            out[label] = e["location"]
            out[f"{label}|設置開始日"] = e["start"]
            out[f"{label}|設置終了日"] = e["end"]
            out[f"{label}|座標"] = e["coord"]
            out[f"{label}|備考"] = e["remark"]
        else:
            out[label] = ""
            out[f"{label}|設置開始日"] = ""
            out[f"{label}|設置終了日"] = ""
            out[f"{label}|座標"] = ""
            out[f"{label}|備考"] = ""

    extra_lines = []
    for j in range(3, len(entries)):
        e = entries[j]
        extra_lines.append(
            f"設置場所{j + 1}: 場所={e['location']}, "
            f"開始={e['start']}, 終了={e['end']}, 座標={e['coord']}, 備考={e['remark']}"
        )
    out["全体備考"] = "\n".join(extra_lines)
    return out


def main():
    schema_obj = json.loads(SRC_JSON.read_text(encoding="utf-8"))
    schema = schema_obj["schema"]
    sheet_name = (schema_obj.get("settings") or {}).get("sheetName") or "Data"

    order = build_order_from_schema(schema)
    matrix, paths = build_header_matrix(order)
    n_cols = len(paths)

    rows, _, _ = read_camera_sheet_with_merges(SRC_XLSX)
    records = collect_records(rows)

    # 出力 workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = sheet_name

    # ヘッダー書き込み
    for r in range(NFB_HEADER_DEPTH):
        for c in range(n_cols):
            v = matrix[r][c]
            if v != "":
                ws_out.cell(row=r + 1, column=c + 1, value=v)

    # 列キー → 列番号 マップ
    col_index: dict[str, int] = {}
    for i, p in enumerate(paths):
        col_index[path_key(p)] = i + 1

    now_ms = int(time.time() * 1000)
    used_ids: set[str] = set()

    for i, rec in enumerate(records):
        row_num = NFB_DATA_START_ROW + i

        rec_id = generate_record_id()
        while rec_id in used_ids:
            rec_id = generate_record_id()
        used_ids.add(rec_id)

        # 固定列
        ws_out.cell(row=row_num, column=col_index["id"], value=rec_id)
        ws_out.cell(row=row_num, column=col_index["No."], value=i + 1)
        ws_out.cell(row=row_num, column=col_index["createdAt"], value=now_ms)
        ws_out.cell(row=row_num, column=col_index["modifiedAt"], value=now_ms)
        # deletedAt 空
        ws_out.cell(row=row_num, column=col_index["createdBy"], value=USER_EMAIL)
        ws_out.cell(row=row_num, column=col_index["modifiedBy"], value=USER_EMAIL)
        # deletedBy 空

        # 動的列
        responses = build_dynamic_responses(rec)
        for k, v in responses.items():
            ck = normalize_header_key(k)
            if ck in col_index and v not in (None, ""):
                ws_out.cell(row=row_num, column=col_index[ck], value=v)

    # 列幅 (見やすさ用、簡易)
    for c in range(1, n_cols + 1):
        ws_out.column_dimensions[get_column_letter(c)].width = 18

    ws_out.freeze_panes = ws_out.cell(row=NFB_DATA_START_ROW, column=1).coordinate

    DST_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(DST_XLSX)

    print(f"OK: wrote {DST_XLSX}")
    print(f"  records: {len(records)}")
    print(f"  columns: {n_cols}")
    print(f"  header path[0..]: {[p[0] for p in paths]}")
    print(f"  header path[1] (where present): {[(p[0], p[1]) for p in paths if len(p) > 1]}")


if __name__ == "__main__":
    main()
