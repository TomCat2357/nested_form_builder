# -*- coding: utf-8 -*-
"""新 7 シート様式から cellmap.gs 起こし用の素材を抽出する開発補助（再実行可）。

対象: form_test/鳥獣保護管理法様式_個人想定.xlsx / _法人想定.xlsx（7 シート）。

色の意味（様式 申請書 L6/L7/L8 の凡例）:
  YEL  FFFFFF00 = 正として吸い取る（authoritative / read as-is）
  PINK FFEAD1DC = 確認用に吸い取る（confirmation; 名簿を集計した数式セル）
  GRN  FF00B050 = 掃き出し場所。吸い取りは行わない（output-only）

出力 (scripts/out/):
  cellmap_seed.tsv : file, sheet, address, color, mergeTL, mergeRange, formula, value, leftLabel, upLabel
                     色付き or 数式のセルのみ（cellmap.gs の転記対象）。色は設計時の典拠で
                     実行時は固定番地を引く。
  merges_new.tsv   : file, sheet, range（結合の左上検証用）

使い方:
  cd gas_for_external_action/choju_intake/scripts
  python extract_cellmap.py            # 既定で form_test の 2 ファイルを処理
  python extract_cellmap.py <xlsx> ... # ファイルを明示指定
"""
import os
import sys
import warnings

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple

warnings.filterwarnings("ignore")  # DataValidation 拡張の警告を抑止

REPO_FORM_TEST = os.path.join(os.path.dirname(__file__), "..", "..", "..", "form_test")
DEFAULT_FILES = [
    os.path.join(REPO_FORM_TEST, "鳥獣保護管理法様式_個人想定.xlsx"),
    os.path.join(REPO_FORM_TEST, "鳥獣保護管理法様式_法人想定.xlsx"),
]

COLOR_NAMES = {"FFFFFF00": "YEL", "FFEAD1DC": "PINK", "FF00B050": "GRN"}


def color_of(cell):
    f = cell.fill
    if f and f.patternType == "solid":
        fg = f.fgColor
        if fg is not None and fg.type == "rgb":
            return COLOR_NAMES.get(fg.rgb, fg.rgb)
    return ""


def clip(s, n=60):
    s = "" if s is None else str(s).replace("\t", " ").replace("\n", "\\n")
    return s if len(s) <= n else s[:n] + "…"


def nearest_left(ws_vals, row, col):
    for c in range(col - 1, 0, -1):
        v = ws_vals.get((row, c))
        if v not in (None, ""):
            return v
    return ""


def nearest_up(ws_vals, row, col):
    for r in range(row - 1, 0, -1):
        v = ws_vals.get((r, col))
        if v not in (None, ""):
            return v
    return ""


def process(path, seed_rows, merge_rows):
    label = os.path.splitext(os.path.basename(path))[0].split("様式_")[-1]  # 個人想定 / 法人想定
    wb_f = openpyxl.load_workbook(path, data_only=False)  # 数式
    wb_v = openpyxl.load_workbook(path, data_only=True)   # キャッシュ値
    for ws in wb_f.worksheets:
        wsv = wb_v[ws.title]
        # 結合: 左上番地 -> レンジ文字列
        merge_tl = {}
        for rng in ws.merged_cells.ranges:
            merge_rows.append((label, ws.title, str(rng)))
            tl = str(rng).split(":")[0]
            merge_tl[tl] = str(rng)
        # ラベル探索用に「値があるセル」を行列インデックス化（数式ではなく表示テキスト優先）
        vals = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is not None and not (isinstance(v, str) and v.startswith("=")):
                    vals[(cell.row, cell.column)] = v
        for row in ws.iter_rows():
            for cell in row:
                color = color_of(cell)
                formula = cell.value if (isinstance(cell.value, str) and cell.value.startswith("=")) else ""
                if not color and not formula:
                    continue
                cached = wsv[cell.coordinate].value
                addr = cell.coordinate
                seed_rows.append((
                    label, ws.title, addr, color,
                    "Y" if addr in merge_tl else "",
                    merge_tl.get(addr, ""),
                    clip(formula, 90),
                    clip(cached, 50),
                    clip(nearest_left(vals, cell.row, cell.column), 40),
                    clip(nearest_up(vals, cell.row, cell.column), 40),
                ))
    print(f"{label}: seed so far {len(seed_rows)}")


def main():
    files = sys.argv[1:] if len(sys.argv) > 1 else DEFAULT_FILES
    out_dir = os.path.join(os.path.dirname(__file__), "out")
    os.makedirs(out_dir, exist_ok=True)

    seed_rows = []
    merge_rows = []
    for p in files:
        process(p, seed_rows, merge_rows)

    def dump(name, header, rows):
        path = os.path.join(out_dir, name)
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write("\t".join(header) + "\n")
            for r in rows:
                f.write("\t".join(str(x) for x in r) + "\n")
        print(f"{name}: {len(rows)} rows -> {path}")

    dump("cellmap_seed.tsv",
         ["file", "sheet", "address", "color", "mergeTL", "mergeRange", "formula", "value", "leftLabel", "upLabel"],
         seed_rows)
    dump("merges_new.tsv", ["file", "sheet", "range"], merge_rows)


if __name__ == "__main__":
    main()
